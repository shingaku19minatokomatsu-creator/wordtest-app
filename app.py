import os
from flask import Flask, request, redirect, session, render_template_string
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook
from pathlib import Path
from reportlab.pdfgen import canvas
from flask import send_file
import io
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import stringWidth
import uuid
from tempfile import gettempdir
import random
import psycopg2
import os
from contextlib import contextmanager
import psycopg2, os
from psycopg2.pool import SimpleConnectionPool
from contextlib import contextmanager
import os
from psycopg2.errors import UniqueViolation



DATABASE_URL = os.environ["DATABASE_URL"]

pool = SimpleConnectionPool(
    minconn=1,
    maxconn=10,
    dsn=DATABASE_URL
)

@contextmanager
def get_db():
    conn = pool.getconn()
    cur = conn.cursor()
    try:
        yield cur
        conn.commit()
    finally:
        cur.close()
        pool.putconn(conn)


# ===== 日本語フォント =====
FONT_PATH = Path("fonts/ipaexm.ttf")
DEFAULT_FONT = "IPAEX_M"

try:
    pdfmetrics.registerFont(TTFont(DEFAULT_FONT, str(FONT_PATH)))
except Exception as e:
    print("⚠ フォント読込失敗 → Helvetica", e)
    DEFAULT_FONT = "Helvetica"



app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret")

EXCEL_PATH = Path("英単語テスト.xlsx")  # ← あなたの単語Excelに合わせてOK

TMPDIR = Path(gettempdir()) / "word_test"
TMPDIR.mkdir(exist_ok=True)


# -------------------------
# HTML（templates無し）
# -------------------------

LOGIN_HTML = """
<!doctype html>
<html lang="ja">
<head>
<meta charset="utf-8">
<title>ログイン</title>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
body{
  font-family: sans-serif;
  background:#f5f5f5;
  margin: 0;
  padding: 16px;
  display:flex;
  justify-content:center;
  align-items:center;
  min-height:100vh;
}

.box{
  background:#fff;
  padding:24px;
  width:100%;
  max-width:360px;   /* ★ 作成画面と同じ思想 */
  border-radius:8px;
  box-shadow: 0 2px 8px rgba(0,0,0,.15);
}

h2{
  text-align:center;
  margin: 0 0 16px;
}

form{
  display: flex;
  flex-direction: column;
  gap: 12px;         /* ★ 縦リズム統一 */
}

input, button{
  width:100%;
  padding:12px;
  font-size:16px;
  box-sizing:border-box;
}

button{
  background:#007bff;
  color:#fff;
  border:none;
  border-radius:6px;
}

a{
  display:block;
  text-align:center;
  margin-top:14px;
}

.error{
  color:red;
  margin-top:10px;
  text-align:center;
}

</style>
</head>
<body>
<div class="box">
<h2>ログイン</h2>

<form method="post">
  <input name="username" placeholder="ID" required>
  <input type="password" name="password" placeholder="パスワード" required>
  <button type="submit">ログイン</button>
</form>

{% if error %}
<div class="error">{{ error }}</div>
{% endif %}

<a href="/register">新規登録</a>
</div>
</body>
</html>
"""


REGISTER_HTML = """
<!doctype html>
<html lang="ja">
<head>
<meta charset="utf-8">
<title>新規登録</title>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
body{font-family:sans-serif;background:#f5f5f5;
display:flex;justify-content:center;align-items:center;height:100vh}
.box{background:#fff;padding:24px;width:320px; max-width:90vw;border-radius:8px;}
input,button{width:100%;padding:12px;font-size:16px;margin-top:10px}
</style>
</head>
<body>
<div class="box">
<h2>新規登録</h2>
<form method="post">
<input name="username" placeholder="名前(ID)" required>
<input type="password" name="password" placeholder="パスワード" required>
<button type="submit">登録</button>
</form>
<a href="/login">戻る</a>
</div>
</body>
</html>
"""

# ===== HTML ======
INDEX_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>単語テスト</title>
<meta name="viewport" content="width=device-width, initial-scale=1.0">

<style>
body {
  font-family: Arial, sans-serif;
  margin: 0 auto;
  padding: 6mm;
  font-size: 14px;
  max-width: none;
  touch-action: pan-y pinch-zoom;
}

html, body {
  overscroll-behavior: none;
}

@media print {

  /* ===== 用紙設定 ===== */
  @page {
    size: A4 landscape;
    margin: 15mm;
  }

  html, body {
    margin: 0;
    padding: 0;
  }

  /* bodyサイズは固定しない */
  body {
    width: auto;
    height: auto;
  }

  /* ===== 印刷対象 ===== */
  .html-test #print-root {
    width: 100%;
    margin: 0;
    padding: 0;
    box-shadow: none;
  }

  /* ===== 行レイアウト（mm固定） ===== */
  .html-test #print-root .item {
    display: grid !important;
    grid-template-columns:
      8mm
      60mm
      32mm
      30mm
      8mm
      60mm
      32mm
      30mm !important;

    height: 9mm !important;
    align-items: center;
    font-size: 10pt;

    break-inside: avoid !important;
    page-break-inside: avoid !important;
  }

  /* ===== canvas を完全制御 ===== */
  .html-test #print-root .item canvas {
    width: 30mm !important;
    height: 7mm !important;
    max-width: 30mm !important;
    max-height: 7mm !important;
    display: block !important;
  }

  /* ===== 余計なUI削除 ===== */
  button,
  .toolbar {
    display: none !important;
  }
}



h2 {
    font-size: 26px;
    margin-bottom: 10px;
}

label {
    display: block;
    font-size: 18px;
    margin-bottom: 4px;
}

input, select, button {
    padding: 12px;
    font-size: 18px;
    width: 100%;
    box-sizing: border-box;
}

.row {
    margin: 15px 0;
}

button {
    background-color: #007bff;
    color: white;
    border: none;
    border-radius: 6px;
    font-size: 20px;
    padding: 14px;
    cursor: pointer;
}

button:hover {
    background-color: #0056c7;
}

.note {
    color: #666;
    font-size: 15px;
    margin-bottom: 10px;
}

/* スマホ用 */
@media (max-width: 600px) {
    body {
        padding: 14px;
        font-size: 17px;
    }
    input, select, button {
        font-size: 18px;
        padding: 14px;
    }
    h2 {
        font-size: 24px;
    }
}
</style>
</head>

<body>

<div style="text-align:right;margin-bottom:10px;">
  <a href="/logout">ログアウト</a>
</div>

<form>
  <div class="row">
    <label>単語帳（シート）</label>
    <select id="sheet">
      {% for s in sheets %}
      <option value="{{s}}">{{s}}</option>
      {% endfor %}
    </select>
  </div>

  <div class="row">
    <label>開始番号</label>
    <input id="start" required>
  </div>

  <div class="row">
    <label>終了番号</label>
    <input id="end" required>
  </div>

  <div class="row">
    <button type="button" onclick="doPdf()">印刷用</button>
  </div>

  <div class="row">
    <button type="button" onclick="doHtml()">テスト</button>
  </div>
</form>

<script>
function getParams(){
  const sheet = document.getElementById('sheet').value;
  const start = document.getElementById('start').value;
  const end   = document.getElementById('end').value;
  if(!sheet || !start || !end){
    alert("シート・開始・終了番号が必要です");
    return null;
  }
  return {sheet, start, end};
}

async function doPdf(){
  const p = getParams();
  if(!p) return;

  const win = window.open("about:blank");

  const res = await fetch("/generate", {
    method: "POST",
    headers: {"Content-Type": "application/json"},
    body: JSON.stringify(p)
  });

  if(!res.ok){
    win.close();
    alert(await res.text());
    return;
  }

  // ★ ここが重要（URLを使わないPDF表示）
  const blob = await res.blob();
  const url = URL.createObjectURL(blob);
  win.location.href = url;
}

async function doHtml(){
  const p = getParams();
  if(!p) return;

  const win = window.open("about:blank");
  const res = await fetch("/generate_html_test", {
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body:JSON.stringify(p)
  });

  if(!res.ok){
    win.close();
    alert(await res.text());
    return;
  }
  const html = await res.text();
  win.document.open();
  win.document.write(html);
  win.document.close();
}
</script>



</body>
</html>
"""


# ①〜④ HTML版テスト機能 追加コード（PDF完全一致レイアウト版）
# 既存 app.py に追記する想定

# ①〜④ HTML版テスト機能 追加コード（PDF完全一致レイアウト版）
# 既存 app.py に追記する想定

HTML_TEST_TEMPLATE = """
<!doctype html>
<html>
<head>
<meta charset=\"utf-8\">
<title>単語テスト（HTML）</title>
<meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">

<style>
/* ===== HTMLテスト画面だけ ===== */
.html-test {
  font-family: Arial, sans-serif;
  background: #f5f5f5;
  margin: 0;
  padding: 16px;
  touch-action: pan-y pinch-zoom;
}

.html-test #print-root {
  background: #fff;
  margin: 0 auto;
  padding: 16px;
  max-width: 1200px;
  box-shadow: 0 2px 8px rgba(0,0,0,.15);
}

/* ===== 画面では canvas をはみ出させない ===== */
.html-test canvas {
  max-width: 100%;
  height: auto;
}


html, body {
  overscroll-behavior: none;
}

/* ===== ヘッダ ===== */
.header {
  display: flex;
  justify-content: space-between;
  margin-bottom: 10mm;
}

/* ===== 画面表示用 ===== */
.item {
  display: grid;
  grid-template-columns:
    44px
    minmax(220px, 1fr)
    minmax(120px, 160px)
    190px
    44px
    minmax(220px, 1fr)
    minmax(120px, 160px)
    190px;

  height: 40px;
  align-items: center;
  font-size: 13px;
  box-sizing: border-box;
}



.answer {
  min-width: 0;
  font-weight: bold;
  color: red;
  opacity: 0.85;

  visibility: hidden;

  font-size: 11px;
  line-height: 1.2;

  white-space: normal;
  word-break: break-word;

  /* ★ここから追加 */
  display: -webkit-box;
  -webkit-line-clamp: 2;
  -webkit-box-orient: vertical;
  overflow: hidden;
}

.answer.show {
  visibility: visible;
}


/* ===== canvas ===== */

canvas {
  display: block;
  background: #f2f2f2;
  border: 1px solid #ccc;

  touch-action: none;     /* ★ canvasだけロック */
  user-select: none;
  pointer-events: auto;
}

.item {
  user-select: none;      /* 文字選択防止だけ */
}

/* ★ item * は消す */

.small-text {
  font-size: 9px;
  line-height: 1.1;
}

/* ===== 印刷時 ===== */
@media print {
  button { display: none; }
}

/* ===== 操作ツールバー ===== */
.toolbar {
  position: fixed;
  top: calc(12px + env(safe-area-inset-top));
  right: calc(12px + env(safe-area-inset-right));
  display: flex;
  gap: 6px;
  background: rgba(255,255,255,0.95);
  padding: 4px;
  border-radius: 8px;
  box-shadow: 0 2px 8px rgba(0,0,0,.2);
  z-index: 1000;
}

.toolbar button {
  font-size: 10px;
  padding: 4px 8px;
  white-space: nowrap; /* 折り返さない */
}


@media (max-width: 900px) {
  .toolbar {
    top: auto;
    bottom: calc(12px + env(safe-area-inset-bottom));
  }
}

@media print {
  .toolbar {
    display: none !important;
  }
}

/* 各問題行を横並びにする */
.word-row {
  display: flex;
  align-items: center;
}

/* 左：問題文 */
.word-row .question {
  flex: 1 1 auto;   /* 残り幅を使う */
  min-width: 0;     /* ★ これが超重要 */
}

/* 右：canvas */
.word-row .answer {
  flex: 0 0 180px;  /* ★ canvas列は固定 */
}



</style>
</head>

<body class="html-test">
<div id="print-root">


<div class=\"header\">
  <div>
    <h2>shingaku19minato test</h2>
    <div>words {{sheet}}（{{start}}～{{end}}）</div>
  </div>
    <div>
        name：
        <canvas width="160" height="28"></canvas><br>
        score：
        <canvas width="160" height="28"></canvas>
    </div>
</div>

<div class="toolbar">
  <button onclick="toggleAll()">解答</button>
  <button onclick="setColor('black')">⚫黒</button>
  <button onclick="setColor('red')">🔴赤</button>
  <button onclick="setMode('eraser')">🧽消</button>
  <button onclick="clearAll()">🗑全消</button>
  <button onclick="window.print()">🖨印刷</button>
</div>


    {% for i in range(20) %}
    {% set item  = items[i] %}
    {% set item2 = items[i+20] %}

    <div class="item">
        <!-- 左（1〜20） -->
        <div>{{item.no}}.</div>
        <div>{{item.q}}</div>
        <div class="answer" id="ans-{{item.no}}">{{item.a}}</div>
        <canvas width="180" height="36"></canvas>

        <!-- 右（21〜40） -->
        <div>{{item2.no}}.</div>
        <div>{{item2.q}}</div>
        <div class="answer" id="ans-{{item2.no}}">{{item2.a}}</div>
        <canvas width="180" height="36"></canvas>
    </div>

    {% endfor %}

<script>
let mode = "pen";
let color = "#000";

function setColor(c){
  color = (c === "red") ? "#d00" : "#000";
  mode = "pen";
}

function setMode(m){
  mode = m;
}

function clearAll(){
  document.querySelectorAll("canvas").forEach(c=>{
    const ctx = c.getContext("2d");
    ctx.save();
    ctx.setTransform(1,0,0,1,0,0);  // ← スケール解除
    ctx.clearRect(0, 0, c.width, c.height);
    ctx.restore();
  });
}


function toggleAll(){
  document.querySelectorAll('.answer')
    .forEach(a => a.classList.toggle('show'));
}


document.querySelectorAll("canvas").forEach(c=>{
  const ratio = window.devicePixelRatio || 1;

  // ===== ① CSSサイズを保存（テンプレそのまま）=====
  const cssW = c.width;
  const cssH = c.height;

  // ===== ② 内部解像度だけ拡大 =====
  c.width  = cssW * ratio;
  c.height = cssH * ratio;

  // ===== ③ 見た目サイズは固定 =====
  c.style.width  = cssW + "px";
  c.style.height = cssH + "px";

  const ctx = c.getContext("2d");

  // ★ ここが最重要（座標系を元に戻す）
  ctx.setTransform(ratio, 0, 0, ratio, 0, 0);


  let drawing = false;

  ctx.lineWidth = 0.6;        // ← 今まで通りでOK
  ctx.lineCap = "round";
  ctx.lineJoin = "round";
  ctx.strokeStyle = color;

  function getPos(e){
    const rect = c.getBoundingClientRect();
    return {
      x: e.clientX - rect.left,
      y: e.clientY - rect.top
    };
  }

  c.addEventListener("touchstart", e=>{
    e.preventDefault();
  }, { passive: false });

  c.addEventListener("pointerdown", e=>{
    e.preventDefault();
    e.stopPropagation();

    drawing = true;
    c.setPointerCapture(e.pointerId);

    const p = getPos(e);
    ctx.beginPath();
    ctx.moveTo(p.x, p.y);
  });

  c.addEventListener("pointermove", e=>{
    if(!drawing) return;
    e.preventDefault();

    const p = getPos(e);

    if(mode === "eraser"){
      ctx.clearRect(p.x - 6, p.y - 6, 12, 12);
    }else{
      ctx.strokeStyle = color;
      ctx.lineTo(p.x, p.y);
      ctx.stroke();
    }
  });

  c.addEventListener("pointerup", e=>{
    drawing = false;
    c.releasePointerCapture(e.pointerId);
  });

  c.addEventListener("pointercancel", ()=>{
    drawing = false;
  });
});


document.querySelectorAll('.answer, .item > div:nth-child(2), .item > div:nth-child(6)')
  .forEach(el=>{
    if(el.textContent.length > 30){
      el.classList.add('small-text');
    }
  });

</script>





</body>
</html>
"""


ADMIN_HTML = """
<!doctype html>
<html lang="ja">
<head>
<meta charset="utf-8">
<title>管理者</title>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
body{
  font-family: sans-serif;
  padding: 16px;
  background: #f5f5f5;
}

h2{
  font-size: 24px;
  margin-bottom: 12px;
}

.controls{
  margin-bottom: 12px;
}

.controls button{
  padding: 8px 12px;
  font-size: 15px;
  margin-right: 6px;
}

/* テーブル全体を包んで横スクロール対応 */
.table-wrap{
  overflow-x: auto;
  background: #fff;
  border-radius: 6px;
}

table{
  width: 100%;
  border-collapse: collapse;
  min-width: 560px;
}

th, td{
  border: 1px solid #ccc;
  padding: 10px;
  font-size: 16px;
  text-align: left;
  white-space: nowrap;
}

th{
  background: #eee;
}

.footer{
  margin-top: 16px;
}

@media (max-width: 600px){
  h2{ font-size: 22px; }
  th, td{ font-size: 15px; padding: 8px; }
}
</style>
</head>

<body>

<h2>管理者画面</h2>

<form method="post" action="/bulk_action">

<div class="controls">
  <button name="action" value="approve">選択承認</button>
  <button name="action" value="delete"
          onclick="return confirm('選択したユーザーを削除しますか？')">
    選択削除
  </button>
  <button type="button" onclick="selectAll()">全選択</button>
  <button type="button" onclick="selectPending()">未承認全選択</button>
</div>

<div class="table-wrap">
<table>
<tr>
  <th></th>
  <th>ID</th>
  <th>名前</th>
  <th>状態</th>
</tr>

{% for u in users %}
<tr>
  <td>
    <input type="checkbox" name="uids" value="{{u[0]}}">
  </td>
  <td>{{u[0]}}</td>
  <td>{{u[1]}}</td>
  <td>{{"OK" if u[2] else "承認待ち"}}</td>
</tr>
{% endfor %}
</table>
</div>

</form>

<div class="footer">
  <a href="/logout">ログアウト</a>
</div>

<script>
function selectAll(){
  document.querySelectorAll("input[name='uids']")
    .forEach(cb => cb.checked = true);
}

function selectPending(){
  document.querySelectorAll("tr").forEach(tr=>{
    if(tr.innerText.includes("承認待ち")){
      const cb = tr.querySelector("input[name='uids']");
      if(cb) cb.checked = true;
    }
  });
}
</script>

</body>
</html>
"""
PENDING_HTML = """
<!doctype html>
<html lang="ja">
<head>
<meta charset="utf-8">
<title>承認待ち</title>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
body{
  font-family:sans-serif;
  background:#f5f5f5;
  display:flex;
  justify-content:center;
  align-items:center;
  height:100vh;
}
.box{
  background:#fff;
  padding:24px;
  width:360px;
  border-radius:8px;
  text-align:center;
}
button{
  margin-top:20px;
  padding:10px;
  width:100%;
  background:#007bff;
  color:#fff;
  border:none;
}
</style>
</head>
<body>
<div class="box">
  <h2>現在、承認待ちです</h2>
  <p>
    このアカウントは、<br>
    管理者による承認後にログインできます。
  </p>

  <form action="/login">
    <button type="submit">ログイン画面に戻る</button>
  </form>
</div>
</body>
</html>
"""




# -------------------------
# ログイン制御
# -------------------------

@app.before_request
def require_login():
    if request.path.startswith(("/login", "/register", "/static", "/favicon.ico")):
        return
    if not session.get("user_id"):
        return redirect("/login")



# -------------------------
# Routes
# -------------------------

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None

    if request.method == "POST":
        u = request.form["username"]
        p = request.form["password"]

        with get_db() as cur:
            cur.execute(
                "SELECT id, username, password_hash, role, approved FROM users WHERE username=%s",
                (u,)
            )
            user = cur.fetchone()

        if not user or not check_password_hash(user[2], p):
            error = "ID または パスワードが違います"
        elif not user[4]:
            return redirect("/pending")

        else:
            session["user_id"] = user[0]
            session["role"] = user[3]
            return redirect("/admin" if user[3] == "admin" else "/")

    return render_template_string(LOGIN_HTML, error=error)




@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        try:
            with get_db() as cur:
                cur.execute(
                    """
                    INSERT INTO users (username, password_hash, role, approved)
                    VALUES (%s, %s, %s, %s)
                    """,
                    (
                        request.form["username"],
                        generate_password_hash(request.form["password"]),
                        "student",
                        False
                    )
                )
        except UniqueViolation:
            return """
            <h3>このIDはすでに使われています</h3>
            <a href="/register">戻る</a>
            """

        return """
        登録しました。承認待ちです。<br>
        <a href='/login'>ログイン画面へ</a>
        """

    return render_template_string(REGISTER_HTML)




@app.route("/")
def index():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    sheets = wb.sheetnames
    return render_template_string(INDEX_HTML, sheets=sheets)


@app.route("/admin")
def admin():
    if session.get("role") != "admin":
        return redirect("/")

    with get_db() as cur:
        cur.execute(
            "SELECT id, username, approved FROM users WHERE role='student'"
        )
        users = cur.fetchall()

    return render_template_string(ADMIN_HTML, users=users)



@app.route("/approve/<int:uid>")
def approve(uid):
    conn, cur = get_db()
    cur.execute("UPDATE users SET approved=true WHERE id=%s", (uid,))
    conn.commit()
    conn.close()
    return redirect("/admin")


@app.route("/reset/<int:uid>")
def reset(uid):
    conn, cur = get_db()
    cur.execute(
        "UPDATE users SET password_hash=%s WHERE id=%s",
        (generate_password_hash("1234"), uid)
    )
    conn.commit()
    conn.close()
    return redirect("/admin")


@app.route("/delete/<int:uid>")
def delete(uid):
    conn, cur = get_db()
    cur.execute("DELETE FROM users WHERE id=%s", (uid,))
    conn.commit()
    conn.close()
    return redirect("/admin")
  
@app.route("/pending")
def pending():
    return render_template_string(PENDING_HTML)



@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")
  
@app.route("/bulk_action", methods=["POST"])
def bulk_action():
    if session.get("role") != "admin":
        return redirect("/")

    action = request.form.get("action")
    uids = request.form.getlist("uids")

    if not uids:
        return redirect("/admin")

    with get_db() as cur:
        if action == "approve":
            cur.executemany(
                "UPDATE users SET approved=true WHERE id=%s",
                [(uid,) for uid in uids]
            )
        elif action == "delete":
            cur.executemany(
                "DELETE FROM users WHERE id=%s",
                [(uid,) for uid in uids]
            )

    return redirect("/admin")

    
@app.route("/generate_html_test", methods=["POST"])
def generate_html_test():
    data = request.get_json()
    sheet = data["sheet"]
    start = int(data["start"])
    end   = int(data["end"])

    rows = load_sheet_rows(EXCEL_PATH, sheet)
    items = pick40(rows, start, end)

    return render_template_string(
        HTML_TEST_TEMPLATE,
        items=items,
        sheet=sheet,
        start=start,
        end=end
    )


  
@app.route("/generate", methods=["POST"])
def generate():
    data = request.get_json()
    sheet = data["sheet"]
    start = int(data["start"])
    end   = int(data["end"])

    rows = load_sheet_rows(EXCEL_PATH, sheet)
    items = pick40(rows, start, end)

    pdf_path = make_two_page_pdf(items, sheet, start, end)

    return send_file(
        str(pdf_path),
        mimetype="application/pdf",
        as_attachment=False
    )

def load_sheet_rows(path, sheet):
    wb = load_workbook(str(path), data_only=True)
    ws = wb[sheet]
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        a, b, c = row
        if a is None and (not b) and (not c):
            continue
        try:
            num = int(float(a))
        except:
            num = None
        rows.append({
            "num": num,
            "q": "" if b is None else str(b),
            "a": "" if c is None else str(c)
        })
    return rows
  


def pick40(rows, start, end):
    r = [x for x in rows if x["num"] is not None and start <= x["num"] <= end]
    random.shuffle(r)
    r = r[:40]
    while len(r) < 40:
        r.append({"num": None, "q": "", "a": ""})
    for i, rr in enumerate(r):
        rr["no"] = i + 1
    return r

def wrap_text(text, font, size, max_width):
    if " " in text:
        units = text.split(" ")
    else:
        units = list(text)

    lines = []
    current = ""

    for u in units:
        test = (current + " " + u).strip() if " " in text else (current + u)
        if stringWidth(test, font, size) <= max_width:
            current = test
        else:
            lines.append(current)
            current = u

    if current:
        lines.append(current)

    return lines

def draw_text_fitted(c, text, font, base_x, base_y, max_width, max_height):
    if not text:
        return

    for size in range(10, 3, -1):
        lines = wrap_text(text, font, size, max_width)
        if len(lines) > 2:
            continue

        total_h = len(lines) * size
        if total_h <= max_height:
            c.setFont(font, size)
            y = base_y
            for ln in lines:
                c.drawString(base_x, y, ln)
                y -= size + 2
            return

def draw_answer_fitted(c, text, font, base_x, base_y, max_width, max_height):
    draw_text_fitted(c, text, font, base_x, base_y, max_width, max_height)


def make_two_page_pdf(items, sheet, start, end):
    filename = TMPDIR / f"{uuid.uuid4().hex}_final.pdf"
    c = canvas.Canvas(str(filename), pagesize=landscape(A4))
    PW, PH = landscape(A4)

    margin = 15*mm
    col_gap = 15*mm
    usable_w = PW - margin*2
    col_w = (usable_w - col_gap)/2

    left_x = margin
    right_x = left_x + col_w + col_gap
    # ====== ページ描画 ======
    def draw_page(mode_label):
        title_y  = PH - 10*mm
        words_y  = title_y - 10*mm
        start_y  = words_y - 14*mm

        c.setFont(DEFAULT_FONT, 16)
        c.drawString(left_x, title_y, "shingaku19minato test")
        
        c.setFont(DEFAULT_FONT, 12)
        c.drawString(left_x, words_y, f"words  {sheet}（{start}～{end}）")
        
        # ←★ これを忘れずに入れる
        c.setFont(DEFAULT_FONT, 12)
        c.drawString(PW - margin - 170, title_y, "name：________________")
        c.drawString(PW - margin - 170, title_y - 8*mm, "score：________________")

        rows_per_col = 20
        bottom = 12*mm
        avail_h = start_y - bottom
        line_h = avail_h / rows_per_col
        if line_h > 12*mm: line_h = 12*mm
        if line_h < 9*mm:  line_h = 9*mm


        # ===== 20行の表を2列に描く =====
        def draw_col(base_x, idx0):
            for i in range(rows_per_col):
                if idx0+i >= len(items): break
                r = items[idx0+i]
        
                y = start_y - i * line_h
        
                # 番号
                c.setFont(DEFAULT_FONT, 10)
                c.drawString(base_x, y, f"{r['no']}.")
        
                # ▼ 幅設定（安全マージン）
                question_width = col_w * 0.50    # 問題の横幅
                answer_width   = col_w * 0.40    # 解答の横幅
                margin_between = col_w * 0.10    # 問題〜解答の間隔
        
                qx = base_x + 10*mm
        
                # ▼ 高さを3行分確保
                max_h = line_h * 3.2
        
                # ▼ 問題
                draw_text_fitted(
                    c, r['q'], DEFAULT_FONT,
                    qx, y,
                    question_width,
                    max_h
                )
        
                if mode_label == "q":
                    lx1 = qx + question_width + 2*mm
                    lx2 = base_x + col_w - 5*mm
                    c.setLineWidth(0.5)
                    c.line(lx1, y - 3, lx2, y - 3)
                else:
                    # ▼ 解答（右に寄せる）
                    ax = base_x + question_width + margin_between
                    # 修正: draw_text_fitted から draw_answer_fitted に変更
                    draw_answer_fitted( 
                        c, r['a'], DEFAULT_FONT,
                        ax, y,
                        answer_width,
                        max_h
                    )



        

        draw_col(left_x, 0)
        draw_col(right_x, 20)

        c.showPage()


    # ===== 1ページ目：問題 =====
    draw_page("q")

    # ===== 2ページ目：解答 =====
    draw_page("a")

    c.save()
    return filename






